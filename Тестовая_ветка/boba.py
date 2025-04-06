import openpyxl

# Загрузим существующий Excel-файл
wb = openpyxl.load_workbook("flatsWithMetrotime1.xlsx")
ws = wb.active  

# Применяем полиномиальную формулу (возведение в квадрат) ко всем значениям в столбце с годом постройки (например, столбец 1)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  # Столбец с годами постройки - 1
    year = row[0].value
    if year is not None:  # Проверяем, что значение года не пустое
        # Рассчитываем x^2 для второго столбца
        year_squared = year ** 2
        
        # Записываем квадрат года в первый столбец
        row[0].value = year_squared
        
        # Добавляем новый столбец для квадрата года
        ws.cell(row=row[0].row, column=2, value=year_squared)

# Сохраняем изменённый файл
wb.save("flats_with_year_squared.xlsx")

print("✅ Годы постройки возведены в квадрат. Сохранено в flats_with_year_squared.xlsx")
